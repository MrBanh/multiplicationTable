#! python3

# multiplicationTable.py - Takes a number n from command line and creates a n*n multiplication table in Excel

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import sys

# Desktop directory
desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop\\')

def multiplicationTable(n):
    try:  
        n = int(n)
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Formatting top row and first column
        boldFont = Font(bold=True)

        # Fill in top row and first column
        for i in range(2, n + 2):
            # row
            sheet.cell(row=1, column=i).value = i - 1
            sheet.cell(row=1, column=i).font = boldFont
            
            # column
            sheet.cell(row=i, column=1).value = i - 1
            sheet.cell(row=i, column=1).font = boldFont


        # Fill in multiplication Table
        for rowNum in range(2, n + 2):
            for colNum in range(2, n + 2):
                sheet.cell(row=rowNum, column = colNum).value =  f'= A{str(rowNum)} * {get_column_letter(colNum)}1'
    except (ValueError, TypeError):
        print('That is not a number. Try again.')
        exit()
    
    except Exception:
        print('Something went wrong. Try again')
        exit()

    # Save excel workbook to desktop
    os.chdir(desktop)
    wb.save('multiplicationTable.xlsx')
    print('Done')

# Implement command line feature
if len(sys.argv) == 1:
    # Print out 'how to use'
    print('\nmultiplicationTable.py - Takes a number n from command line and creates a n*n multiplication table in Excel')
    print('\nHow to Use via Command Line:')
    print('\n\tmultiplicationTable <number>\n')
else:
    multiplicationTable(sys.argv[1])